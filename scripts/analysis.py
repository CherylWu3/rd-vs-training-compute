"""
R&D Compute vs. Final Training Run Compute for Chinese AI Labs
==============================================================

Motivation: Most of a lab's R&D compute spending goes to experiments,
derisking runs, synthetic data generation, etc. — NOT to final training runs.
We showed this for OpenAI in 2024 (https://epoch.ai/data-insights/openai-compute-spend).

This script does the same analysis for Minimax and Zhipu, using:
  - R&D compute spending from IPO filings (known)
  - Final training run compute estimates from Epoch (to be filled in)

Authors: JS, Cheryl Wu
"""

from __future__ import annotations

import os
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd
import openpyxl
from dataclasses import dataclass

_SCRIPT_DIR = os.path.dirname(__file__)
_DATA_DIR = os.path.join(_SCRIPT_DIR, "..", "data")
_OUTPUT_DIR = os.path.join(_SCRIPT_DIR, "..", "output")

# RMB/USD conversion rate (2025 H1 avg market rate), used by load_zhipu_financials.
RMB_PER_USD = 7.2


# =============================================================================
# LOAD FINANCIAL DATA FROM EXCEL (12-month spending windows)
# =============================================================================

def load_minimax_financials(path=os.path.join(_DATA_DIR, "hk_ipo_financials_one_sheet.xlsx")):
    """Load MiniMax compute spending for Q4 2024 + Q1-Q3 2025 (12 months).

    Computes Q4 2024 = FY2024 - 9M2024, then adds 9M2025.
    All values in USD millions (native unit of the Excel file).
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    data = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        metric = row[0]
        if metric:
            # Columns: metric, unit, 2022, 2023, 2024, 2024_9M, 2025_9M, source, definition
            data[metric] = {
                "2024": row[4] if row[4] is not None else 0,
                "2024_9M": row[5] if row[5] is not None else 0,
                "2025_9M": row[6] if row[6] is not None else 0,
            }

    # Q4 2024 = Full Year 2024 minus first 9 months of 2024
    rd_q4 = data["Compute spending (R&D)"]["2024"] - data["Compute spending (R&D)"]["2024_9M"]
    inf_q4 = (data["Compute spending (Operating cost / inference cloud)"]["2024"]
              - data["Compute spending (Operating cost / inference cloud)"]["2024_9M"])
    tot_q4 = (data["Compute spending (Total disclosed)"]["2024"]
              - data["Compute spending (Total disclosed)"]["2024_9M"])

    return {
        "rd_compute": rd_q4 + data["Compute spending (R&D)"]["2025_9M"],
        "inference_compute": inf_q4 + data["Compute spending (Operating cost / inference cloud)"]["2025_9M"],
        "total_compute": tot_q4 + data["Compute spending (Total disclosed)"]["2025_9M"],
        "unit": "USD mn",
        "window": "Q4 2024 – Q3 2025 (12 months)",
    }


def load_zhipu_financials(path=os.path.join(_DATA_DIR, "zhipu_ipo_financial_metrics.xlsx")):
    """Load Zhipu compute spending for H2 2024 + H1 2025 (12 months).

    Computes H2 2024 = FY2024 - H1 2024, then adds H1 2025.
    Native unit is RMB millions; converted to USD at 7.2 RMB/USD
    (approximate 2025 H1 average market rate).
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    data = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        metric = row[0]
        if metric:
            # Columns: metric, unit, 2022, 2023, 2024, 2024H1, 2025H1, source, definition
            data[metric] = {
                "2024": row[4] if row[4] is not None else 0,
                "2024H1": row[5] if row[5] is not None else 0,
                "2025H1": row[6] if row[6] is not None else 0,
            }

    # H2 2024 = Full Year 2024 minus H1 2024
    rd_h2 = data["Compute spending (R&D)"]["2024"] - data["Compute spending (R&D)"]["2024H1"]
    nonrd_h2 = (data["Compute spending (Non-R&D subtotal)"]["2024"]
                - data["Compute spending (Non-R&D subtotal)"]["2024H1"])
    tot_h2 = (data["Compute spending (Total disclosed)"]["2024"]
              - data["Compute spending (Total disclosed)"]["2024H1"])

    rd_rmb = rd_h2 + data["Compute spending (R&D)"]["2025H1"]
    nonrd_rmb = nonrd_h2 + data["Compute spending (Non-R&D subtotal)"]["2025H1"]
    tot_rmb = tot_h2 + data["Compute spending (Total disclosed)"]["2025H1"]

    return {
        "rd_compute": rd_rmb / RMB_PER_USD,
        "inference_compute": nonrd_rmb / RMB_PER_USD,
        "total_compute": tot_rmb / RMB_PER_USD,
        "unit": "USD mn (converted from RMB)",
        "window": "H2 2024 – H1 2025 (12 months)",
        "rmb_per_usd": RMB_PER_USD,
    }


# =============================================================================
# FLOP-TO-COST CONVERSION
# =============================================================================
# We convert training FLOP to cloud cost via:
#   cost = (flop / (gpu_peak_flops * mfu)) / 3600 * cost_per_gpu_hour
#
# Uncertainty ranges:
#   MFU:    0.15 (pessimistic) to 0.35 (optimistic)
#   Price:  $1.50 to $3.00 per H800-hour in China
#
# The low MFU of 0.15 accounts for: communication overhead at scale,
# pipeline bubble, checkpointing, restarts, and smaller/less-optimized runs.
# The high MFU of 0.35 is achievable for well-optimized large pretraining.

GPU_PEAK_FLOPS = 9.89e14  # FLOP/s (H800 dense BF16/FP16)
MFU_LOW = 0.15    # pessimistic -> more GPU-hours needed -> higher cost
MFU_HIGH = 0.35   # optimistic -> fewer GPU-hours -> lower cost
COST_PER_GPU_HOUR_LOW = 1.50   # cheap end of China cloud pricing
COST_PER_GPU_HOUR_HIGH = 3.00  # expensive end


@dataclass
class TrainingRunEstimate:
    """Estimate of the MARGINAL compute for a single final training run.

    'Marginal' means: for finetuned/post-trained models, we count only the
    incremental compute (finetune/RL), NOT the base model pretraining (which
    is counted separately under the base model entry).
    """
    model_name: str
    org: str  # "minimax" or "zhipu"
    release_date: str  # approximate, YYYY-MM format
    # Marginal training FLOP — use finetune FLOP for post-trained models,
    # full training FLOP for pretraining runs.
    # Treated as p10 and p90 of the FLOP distribution.
    training_flop_low: float | None = None
    training_flop_high: float | None = None
    # If True, FLOP estimate is from Epoch's dataset (higher confidence).
    # If False, it's our rough order-of-magnitude guess (flagged as TODO).
    epoch_estimate: bool = False
    # Per-model MFU override (10th/90th percentiles). If set, uses this
    # instead of the global MFU. Use lower MFU for RL/inference-heavy
    # workloads (rollout generation has much worse utilization than pretraining).
    mfu_override_low: float | None = None
    mfu_override_high: float | None = None
    # Per-model GPU peak FLOP/s override. If set, uses this instead of the
    # global GPU_PEAK_FLOPS. Use for models trained in FP8 (9.89e14).
    gpu_peak_flops_override: float | None = None
    # How to sample FLOP: "lognormal" (default) or "uniform".
    # Use "uniform" for regression-derived bounds (p5/p95 from bootstrap +
    # uniform param draws), since the original distribution was bounded.
    flop_distribution: str = "lognormal"
    # --- For derived finetune models ---
    # When set, FLOP is drawn in the MC as: base_flop_sample * fraction_sample,
    # properly propagating uncertainty from both the base model FLOP and the
    # finetune fraction. training_flop_low/high are then display-only (extreme
    # corners of the product).
    # base_model_ref: model_name of a model earlier in the same list, whose
    #   drawn FLOP samples will be reused (correlated draws).
    # base_flop_low/high: if base is NOT in the model list, specify its bounds.
    base_model_ref: str | None = None
    base_flop_low: float | None = None
    base_flop_high: float | None = None
    base_flop_distribution: str = "lognormal"
    finetune_fraction_low: float | None = None   # p10 of fraction (e.g. 0.01)
    finetune_fraction_high: float | None = None  # p90 of fraction (e.g. 0.10)
    notes: str = ""


# =============================================================================
# REGRESSION-BASED FLOP ESTIMATES (from external scripts)
# =============================================================================
# p5 / p50 / p95 values (90% CI) computed by running the regression scripts
# with bootstrap (N=20,000, seed=42) + MC over unknown parameters.

from video_regression import get_video_percentiles
from image_regression import get_image_percentiles
from speech_regression import get_speech_percentiles

_video_pcts = get_video_percentiles()
_image_pcts = get_image_percentiles()
_speech_pcts = get_speech_percentiles()

# --- From video_regression.py ---
HAILUO02_FLOP_P5 = _video_pcts["hailuo02"]["p5"]
HAILUO02_FLOP_P50 = _video_pcts["hailuo02"]["p50"]
HAILUO02_FLOP_P95 = _video_pcts["hailuo02"]["p95"]

COGVIDEOX_FLOP_P5 = _video_pcts["cogvideox"]["p5"]
COGVIDEOX_FLOP_P50 = _video_pcts["cogvideox"]["p50"]
COGVIDEOX_FLOP_P95 = _video_pcts["cogvideox"]["p95"]

# Hailuo Video-01: needed as base for director-control finetune estimates
HAILUO01_FLOP_P5 = _video_pcts["hailuo01"]["p5"]
HAILUO01_FLOP_P50 = _video_pcts["hailuo01"]["p50"]
HAILUO01_FLOP_P95 = _video_pcts["hailuo01"]["p95"]

# --- From image_regression.py ---
IMAGE01_FLOP_P5 = _image_pcts["image01"]["p5"]
IMAGE01_FLOP_P50 = _image_pcts["image01"]["p50"]
IMAGE01_FLOP_P95 = _image_pcts["image01"]["p95"]

# --- From speech_regression.py ---
SPEECH01_HD_FLOP_P5 = _speech_pcts["speech01_hd"]["p5"]
SPEECH01_HD_FLOP_P50 = _speech_pcts["speech01_hd"]["p50"]
SPEECH01_HD_FLOP_P95 = _speech_pcts["speech01_hd"]["p95"]

SPEECH01_TURBO_FLOP_P5 = _speech_pcts["speech01_turbo"]["p5"]
SPEECH01_TURBO_FLOP_P50 = _speech_pcts["speech01_turbo"]["p50"]
SPEECH01_TURBO_FLOP_P95 = _speech_pcts["speech01_turbo"]["p95"]

SPEECH02_HD_FLOP_P5 = _speech_pcts["speech02_hd"]["p5"]
SPEECH02_HD_FLOP_P50 = _speech_pcts["speech02_hd"]["p50"]
SPEECH02_HD_FLOP_P95 = _speech_pcts["speech02_hd"]["p95"]

SPEECH02_TURBO_FLOP_P5 = _speech_pcts["speech02_turbo"]["p5"]
SPEECH02_TURBO_FLOP_P50 = _speech_pcts["speech02_turbo"]["p50"]
SPEECH02_TURBO_FLOP_P95 = _speech_pcts["speech02_turbo"]["p95"]

# --- MiniMax-M2 pretraining FLOP bounds (pure pretraining, no RL overhead) ---
_M2_PRETRAIN_FLOP_LOW = 6e23    # 10B active x 10T tokens x 6
_M2_PRETRAIN_FLOP_HIGH = 1.5e24  # 10B active x 25T tokens x 6


# =============================================================================
# MODEL-LEVEL TRAINING ESTIMATES
# =============================================================================
# MiniMax: models released in calendar year 2025.
# Spending window: Q4 2024 through Q3 2025 (12 months).
#
# Zhipu: models released Q4 2024 through Q3 2025.
# Spending window: H2 2024 + H1 2025 (12 months).
#
# For models with Epoch FLOP estimates, epoch_estimate=True.
# For models WITHOUT Epoch estimates, we use regression-based estimates or
# manual bounds, with epoch_estimate=False.
#
# IMPORTANT: We use MARGINAL compute — for models fine-tuned on a base,
# we use the finetune/post-training FLOP only (from Epoch's "Finetune
# compute" column), not the total which includes the base pretraining.

MINIMAX_MODELS = [
    # ---- LLMs with Epoch estimates ----
    TrainingRunEstimate(
        model_name="MiniMax-Text-01",
        org="minimax",
        release_date="2025-01",
        training_flop_low=3.142e24,
        training_flop_high=3.142e24,
        epoch_estimate=True,
        notes="456B MoE, 45.9B active params, 11.4T tokens. Pretraining run. "
              "Epoch: 6*45.9e9*1.14e13 ≈ 3.14e24. Hardware: H800.",
    ),
    TrainingRunEstimate(
        model_name="MiniMax-VL-01",
        org="minimax",
        release_date="2025-01",
        training_flop_low=1.41e23,   # marginal: finetune compute from Epoch
        training_flop_high=1.41e23,
        epoch_estimate=True,
        notes="Vision-language model. Finetuned on Text-01 base. "
              "Epoch finetune FLOP=1.41e23 (marginal only).",
    ),
    # M1 finetune has two phases (from Epoch's finetune compute notes):
    #   1. Continual pretraining: 7.5T tokens * 45.9B active * 6 = 2.0655e24 FLOP
    #   2. RL: 3 weeks on 512 H800s at ~0.3 MFU = 2.756e23 FLOP (M1-80k)
    #          1.5 weeks (M1-40k, stopped midway) = 1.378e23 FLOP
    # We split these into separate entries with appropriate MFU.
    TrainingRunEstimate(
        model_name="MiniMax-M1 (continual pretrain)",
        org="minimax",
        release_date="2025-06",
        training_flop_low=2.0655e24,  # same for both M1-40k and M1-80k
        training_flop_high=2.0655e24,
        epoch_estimate=True,
        # Uses global MFU (0.15-0.35) — this is standard pretraining
        notes="Continual pretraining phase of M1: 7.5T tokens at 45.9B active params. "
              "Epoch: 6*45.9e9*7.5e12 = 2.0655e24. Same for M1-40k and M1-80k. "
              "Uses global MFU (standard pretraining workload).",
    ),
    TrainingRunEstimate(
        model_name="MiniMax-M1-80k (RL phase)",
        org="minimax",
        release_date="2025-06",
        training_flop_low=2.756e23,   # M1-80k RL (full 3 weeks)
        training_flop_high=2.756e23,  # M1-80k RL (full 3 weeks)
        epoch_estimate=True,
        mfu_override_low=0.01,  # RL is mostly inference (rollouts) -> very low MFU
        mfu_override_high=0.10,
        notes="RL phase of M1-80k: 3 weeks on 512 H800s. "
              "Epoch assumes 0.3 MFU for FLOP calculation, but we use 0.01-0.10 for "
              "cost estimation (RL workload is inference-heavy). "
              "SFT stage assumed negligible per Epoch.",
    ),
    TrainingRunEstimate(
        model_name="MiniMax-M1-40k (RL phase)",
        org="minimax",
        release_date="2025-06",
        training_flop_low=1.378e23,   # M1-40k RL (~1.5 weeks, stopped midway)
        training_flop_high=1.378e23,  # M1-40k RL (~1.5 weeks)
        epoch_estimate=True,
        mfu_override_low=0.01,  # RL is mostly inference (rollouts) -> very low MFU
        mfu_override_high=0.10,
        notes="RL phase of M1-40k: ~1.5 weeks on 512 H800s (stopped midway through "
              "the M1-80k 3-week run). Both M1-40k and M1-80k were released. "
              "Epoch assumes 0.3 MFU for FLOP calculation, but we use 0.01-0.10 for "
              "cost estimation (RL workload is inference-heavy). "
              "SFT stage assumed negligible per Epoch.",
    ),
    # ---- Models WITHOUT Epoch FLOP estimates ----
    TrainingRunEstimate(
        model_name="MiniMax-M2 (pretraining)",
        org="minimax",
        release_date="2025-10",
        training_flop_low=_M2_PRETRAIN_FLOP_LOW,   # 10B active x 10T tokens x 6
        training_flop_high=_M2_PRETRAIN_FLOP_HIGH,  # 10B active x 25T tokens x 6
        epoch_estimate=False,
        gpu_peak_flops_override=1.513e15,  # H800 FP8 (confirmed FP8 training)
        flop_distribution="uniform",
        notes="NO EPOCH ESTIMATE. 230B total / 10B active MoE (per HF page). "
              "Pure pretraining FLOP only (RL accounted separately). "
              "Low: 10T tokens. High: 25T tokens (90th pctl). "
              "Oct 2025 release — training likely Q2-Q3 2025. "
              "Trained in FP8 (confirmed by MiniMax engineer on HuggingFace).",
    ),
    TrainingRunEstimate(
        model_name="MiniMax-M2 (RL phase)",
        org="minimax",
        release_date="2025-10",
        training_flop_low=6e21,   # display: base_low * frac_low
        training_flop_high=1.5e23,  # display: base_high * frac_high
        epoch_estimate=False,
        mfu_override_low=0.01,   # RL is inference-heavy -> very low MFU
        mfu_override_high=0.10,
        gpu_peak_flops_override=1.513e15,  # H800 FP8
        base_model_ref="MiniMax-M2 (pretraining)",  # correlated with pretrain draw
        finetune_fraction_low=0.01,   # 1% of base
        finetune_fraction_high=0.10,  # 10% of base
        notes="RL phase of M2. FLOP drawn as base_flop * LogNormal(p10=1%,p90=10%) in MC. "
              "MFU 0.01-0.10 (RL workload is inference-heavy, same as M1 RL).",
    ),
    TrainingRunEstimate(
        model_name="MiniMax-M2.1",
        org="minimax",
        release_date="2025-12",
        training_flop_low=6e21,     # display: base_low * frac_low
        training_flop_high=1.5e23,  # display: base_high * frac_high
        epoch_estimate=False,
        mfu_override_low=0.01,   # RL is inference-heavy -> very low MFU
        mfu_override_high=0.10,
        base_model_ref="MiniMax-M2 (pretraining)",  # correlated with pretrain draw
        finetune_fraction_low=0.01,   # 1% of base
        finetune_fraction_high=0.10,  # 10% of base
        notes="Reasoning model update to M2 (229B MoE, 10B active). "
              "RL finetune on M2. FLOP drawn as base_flop * LogNormal(p10=1%,p90=10%) in MC. "
              "LOW MFU: RL workload is inference-heavy.",
    ),
    TrainingRunEstimate(
        model_name="Hailuo-02 (pretraining)",
        org="minimax",
        release_date="2025-06",
        # From video_regression.py bootstrap + MC prediction
        training_flop_low=HAILUO02_FLOP_P5,
        training_flop_high=HAILUO02_FLOP_P95,
        epoch_estimate=False,
        flop_distribution="uniform",
        notes="Video generation model. FLOP from video_regression.py "
              "(bootstrap N=20000 + MC over params U(15B,30B), 1080p). "
              "Next-gen video model, successor to Hailuo Video-01.",
    ),
    TrainingRunEstimate(
        model_name="Hailuo-02 (SFT phase)",
        org="minimax",
        release_date="2025-06",
        training_flop_low=HAILUO02_FLOP_P5 * 0.001,
        training_flop_high=HAILUO02_FLOP_P95 * 0.05,
        epoch_estimate=False,
        base_model_ref="Hailuo-02 (pretraining)",  # correlated with pretrain draw
        finetune_fraction_low=0.001,   # 0.1% of base
        finetune_fraction_high=0.05,   # 5% of base
        notes="SFT/post-training phase of Hailuo-02. FLOP drawn as "
              "base_flop * LogNormal(p10=0.1%,p90=5%) in MC. Global MFU (standard finetuning).",
    ),
    TrainingRunEstimate(
        model_name="Hailuo I2V-01-Director",
        org="minimax",
        release_date="2025-03",
        training_flop_low=HAILUO01_FLOP_P5 * 0.001,
        training_flop_high=HAILUO01_FLOP_P95 * 0.05,
        epoch_estimate=False,
        # Hailuo Video-01 is NOT in model list; provide base bounds directly.
        # base_model_ref ensures both I2V and T2V share the same base draw.
        base_model_ref="Hailuo Video-01",
        base_flop_low=HAILUO01_FLOP_P5,
        base_flop_high=HAILUO01_FLOP_P95,
        base_flop_distribution="uniform",    # from regression
        finetune_fraction_low=0.001,   # 0.1% of base
        finetune_fraction_high=0.05,   # 5% of base
        notes="Image-to-video director-control variant. Finetune on "
              "Hailuo Video-01. FLOP drawn as base_flop * LogNormal(p10=0.1%,p90=5%) in MC.",
    ),
    TrainingRunEstimate(
        model_name="Hailuo T2V-01-Director",
        org="minimax",
        release_date="2025-03",
        training_flop_low=HAILUO01_FLOP_P5 * 0.001,
        training_flop_high=HAILUO01_FLOP_P95 * 0.05,
        epoch_estimate=False,
        base_model_ref="Hailuo Video-01",
        base_flop_low=HAILUO01_FLOP_P5,
        base_flop_high=HAILUO01_FLOP_P95,
        base_flop_distribution="uniform",
        finetune_fraction_low=0.001,
        finetune_fraction_high=0.05,
        notes="Text-to-video director-control variant. Finetune on "
              "Hailuo Video-01. FLOP drawn as base_flop * LogNormal(p10=0.1%,p90=5%) in MC.",
    ),
    TrainingRunEstimate(
        model_name="Image-01 (pretraining)",
        org="minimax",
        release_date="2025-02",
        # From image_regression.py bootstrap + MC prediction
        training_flop_low=IMAGE01_FLOP_P5,
        training_flop_high=IMAGE01_FLOP_P95,
        epoch_estimate=False,
        flop_distribution="uniform",
        notes="Text-to-image model. FLOP from image_regression.py "
              "(bootstrap N=20000 + MC over params U(6B,18B), 2048x2048).",
    ),
    TrainingRunEstimate(
        model_name="Image-01 (SFT phase)",
        org="minimax",
        release_date="2025-02",
        training_flop_low=IMAGE01_FLOP_P5 * 0.001,
        training_flop_high=IMAGE01_FLOP_P95 * 0.05,
        epoch_estimate=False,
        base_model_ref="Image-01 (pretraining)",
        finetune_fraction_low=0.001,
        finetune_fraction_high=0.05,
        notes="SFT/post-training phase of Image-01. FLOP drawn as "
              "base_flop * LogNormal(p10=0.1%,p90=5%) in MC. Global MFU (standard finetuning).",
    ),
    TrainingRunEstimate(
        model_name="MiniMax Speech-01-HD (pretraining)",
        org="minimax",
        release_date="2025-01",
        # From speech_regression.py bootstrap + MC prediction
        training_flop_low=SPEECH01_HD_FLOP_P5,
        training_flop_high=SPEECH01_HD_FLOP_P95,
        epoch_estimate=False,
        flop_distribution="uniform",
        notes="TTS model. FLOP from speech_regression.py "
              "(bootstrap N=20000 + MC over params U(0.3B,2.5B), 48kHz).",
    ),
    TrainingRunEstimate(
        model_name="MiniMax Speech-01-HD (SFT phase)",
        org="minimax",
        release_date="2025-01",
        training_flop_low=SPEECH01_HD_FLOP_P5 * 0.001,
        training_flop_high=SPEECH01_HD_FLOP_P95 * 0.05,
        epoch_estimate=False,
        base_model_ref="MiniMax Speech-01-HD (pretraining)",
        finetune_fraction_low=0.001,
        finetune_fraction_high=0.05,
        notes="SFT/post-training phase of Speech-01-HD. FLOP drawn as "
              "base_flop * LogNormal(p10=0.1%,p90=5%) in MC. Global MFU (standard finetuning).",
    ),
    TrainingRunEstimate(
        model_name="MiniMax Speech-01-turbo (pretraining)",
        org="minimax",
        release_date="2025-01",
        # From speech_regression.py bootstrap + MC prediction
        training_flop_low=SPEECH01_TURBO_FLOP_P5,
        training_flop_high=SPEECH01_TURBO_FLOP_P95,
        epoch_estimate=False,
        flop_distribution="uniform",
        notes="TTS model (turbo variant). FLOP from speech_regression.py "
              "(bootstrap N=20000 + MC over params U(0.4B,1.5B), 44.1kHz).",
    ),
    TrainingRunEstimate(
        model_name="MiniMax Speech-01-turbo (SFT phase)",
        org="minimax",
        release_date="2025-01",
        training_flop_low=SPEECH01_TURBO_FLOP_P5 * 0.001,
        training_flop_high=SPEECH01_TURBO_FLOP_P95 * 0.05,
        epoch_estimate=False,
        base_model_ref="MiniMax Speech-01-turbo (pretraining)",
        finetune_fraction_low=0.001,
        finetune_fraction_high=0.05,
        notes="SFT/post-training phase of Speech-01-turbo. FLOP drawn as "
              "base_flop * LogNormal(p10=0.1%,p90=5%) in MC. Global MFU (standard finetuning).",
    ),
    TrainingRunEstimate(
        model_name="MiniMax-Speech-02-HD (pretraining)",
        org="minimax",
        release_date="2025-05",
        # From speech_regression.py bootstrap + MC prediction
        training_flop_low=SPEECH02_HD_FLOP_P5,
        training_flop_high=SPEECH02_HD_FLOP_P95,
        epoch_estimate=False,
        flop_distribution="uniform",
        notes="TTS model. FLOP from speech_regression.py "
              "(bootstrap N=20000 + MC over params U(0.5B,3B), 44.1kHz).",
    ),
    TrainingRunEstimate(
        model_name="MiniMax-Speech-02-HD (SFT phase)",
        org="minimax",
        release_date="2025-05",
        training_flop_low=SPEECH02_HD_FLOP_P5 * 0.001,
        training_flop_high=SPEECH02_HD_FLOP_P95 * 0.05,
        epoch_estimate=False,
        base_model_ref="MiniMax-Speech-02-HD (pretraining)",
        finetune_fraction_low=0.001,
        finetune_fraction_high=0.05,
        notes="SFT/post-training phase of Speech-02-HD. FLOP drawn as "
              "base_flop * LogNormal(p10=0.1%,p90=5%) in MC. Global MFU (standard finetuning).",
    ),
    TrainingRunEstimate(
        model_name="MiniMax-Speech-02-turbo (pretraining)",
        org="minimax",
        release_date="2025-05",
        # From speech_regression.py bootstrap + MC prediction
        training_flop_low=SPEECH02_TURBO_FLOP_P5,
        training_flop_high=SPEECH02_TURBO_FLOP_P95,
        epoch_estimate=False,
        flop_distribution="uniform",
        notes="TTS model (turbo variant). FLOP from speech_regression.py "
              "(bootstrap N=20000 + MC over params U(0.6B,2B), 44.1kHz).",
    ),
    TrainingRunEstimate(
        model_name="MiniMax-Speech-02-turbo (SFT phase)",
        org="minimax",
        release_date="2025-05",
        training_flop_low=SPEECH02_TURBO_FLOP_P5 * 0.001,
        training_flop_high=SPEECH02_TURBO_FLOP_P95 * 0.05,
        epoch_estimate=False,
        base_model_ref="MiniMax-Speech-02-turbo (pretraining)",
        finetune_fraction_low=0.001,
        finetune_fraction_high=0.05,
        notes="SFT/post-training phase of Speech-02-turbo. FLOP drawn as "
              "base_flop * LogNormal(p10=0.1%,p90=5%) in MC. Global MFU (standard finetuning).",
    ),
]

ZHIPU_MODELS = [
    # ---- LLMs with Epoch estimates ----
    TrainingRunEstimate(
        model_name="GLM-4.5",
        org="zhipu",
        release_date="2025-08",
        training_flop_low=4.42e24,
        training_flop_high=4.42e24,
        epoch_estimate=True,
        notes="Core LLM pretraining. 32B active params (MoE), ~23T tokens. "
              "Epoch: 6*32e9*23e12 = 4.42e24. Released Aug 2025 — after H1 "
              "reporting period but training likely started in H1.",
    ),
    TrainingRunEstimate(
        model_name="GLM-4.5-Air",
        org="zhipu",
        release_date="2025-08",
        training_flop_low=1.656e24,
        training_flop_high=1.656e24,
        epoch_estimate=True,
        notes="Smaller variant. 12B active params, ~23T tokens. "
              "Epoch: 6*12e9*23e12 = 1.656e24. Same timing caveat as GLM-4.5.",
    ),
    TrainingRunEstimate(
        model_name="GLM-4-32B-0414",
        org="zhipu",
        release_date="2025-04",
        training_flop_low=2.88e24,
        training_flop_high=2.88e24,
        epoch_estimate=True,
        notes="32B dense model. Epoch: 6*32e9*15e12 = 2.88e24. "
              "Within H1 reporting period.",
    ),
    TrainingRunEstimate(
        model_name="GLM-4-9B-0414",
        org="zhipu",
        release_date="2025-04",
        training_flop_low=8.10e23,
        training_flop_high=8.10e23,
        epoch_estimate=True,
        notes="9B dense model. Epoch: 6*9e9*15e12 = 8.10e23. "
              "Within H1 reporting period.",
    ),
    TrainingRunEstimate(
        model_name="GLM-Z1-Rumination-32B-0414",
        org="zhipu",
        release_date="2025-04",
        # Epoch lists total=2.88e24 (same as GLM-4-32B), no finetune column.
        # This is likely the same pretraining as GLM-4-32B plus RL, but the
        # RL portion isn't separated out. To avoid double-counting the base
        # pretraining (already counted under GLM-4-32B-0414), we estimate
        # the marginal RL cost as 1-10% of base pretraining.
        epoch_estimate=False,  # the marginal estimate is ours, not Epoch's
        mfu_override_low=0.01,  # RL is mostly inference (rollouts) -> very low MFU
        mfu_override_high=0.10,
        base_model_ref="GLM-4-32B-0414",  # correlated with base pretrain draw
        finetune_fraction_low=0.01,   # 1% of base
        finetune_fraction_high=0.10,  # 10% of base
        notes="NO DIRECT EPOCH ESTIMATE FOR MARGINAL COST. "
              "Epoch total=2.88e24, same as GLM-4-32B base. "
              "Likely shares base pretraining with GLM-4-32B; marginal cost "
              "is the RL/rumination post-training. FLOP drawn as "
              "base_flop * LogNormal(p10=1%,p90=10%) in MC. "
              "LOW MFU: RL workload is inference-heavy.",
    ),
    TrainingRunEstimate(
        model_name="GLM-4.5V",
        org="zhipu",
        release_date="2025-08",
        training_flop_low=1.44e23,   # marginal: finetune compute from Epoch
        training_flop_high=1.44e23,
        epoch_estimate=True,
        notes="Vision-language model. Finetuned on GLM-4.5-Air base. "
              "Epoch finetune FLOP=1.44e23 (marginal only).",
    ),
    TrainingRunEstimate(
        model_name="GLM-4.1V-Thinking",
        org="zhipu",
        release_date="2025-08",
        training_flop_low=1.08e23,   # marginal: finetune compute from Epoch
        training_flop_high=1.08e23,
        epoch_estimate=True,
        notes="9B vision-language reasoning model. Finetuned on GLM-4-9B-0414 base. "
              "Epoch finetune FLOP=1.08e23 (marginal only). "
              "Uses Reinforcement Learning with Curriculum Sampling (RLCS).",
    ),
    TrainingRunEstimate(
        model_name="GLM-4.6",
        org="zhipu",
        release_date="2025-09",
        # Epoch lists 4.42e24 total (same 6ND as GLM-4.5), but this is
        # a post-training update on GLM-4.5 — only count marginal cost.
        epoch_estimate=False,
        mfu_override_low=0.01,   # RL is mostly inference (rollouts) -> very low MFU
        mfu_override_high=0.10,
        base_model_ref="GLM-4.5",  # correlated with GLM-4.5 pretrain draw
        finetune_fraction_low=0.01,   # 1% of base
        finetune_fraction_high=0.10,  # 10% of base
        notes="Post-training update on GLM-4.5 (same architecture, same 23T tokens). "
              "Epoch total=4.42e24 but that double-counts base pretraining. "
              "Marginal cost is the RL post-training. FLOP drawn as "
              "base_flop * LogNormal(p10=1%,p90=10%) in MC. "
              "LOW MFU: RL workload is inference-heavy.",
    ),
    # ---- Models with regression-based or manual estimates ----
    TrainingRunEstimate(
        model_name="CogVideoX",
        org="zhipu",
        release_date="2024-10",
        # From video_regression.py bootstrap prediction
        training_flop_low=COGVIDEOX_FLOP_P5,
        training_flop_high=COGVIDEOX_FLOP_P95,
        epoch_estimate=False,
        flop_distribution="uniform",
        notes="Video generation model. FLOP from video_regression.py "
              "(bootstrap N=20000, params=5B exactly known, 1360x768).",
    ),
    TrainingRunEstimate(
        model_name="CogVideoX (SFT phase)",
        org="zhipu",
        release_date="2024-10",
        epoch_estimate=False,
        base_model_ref="CogVideoX",
        finetune_fraction_low=0.001,   # 0.1% of base
        finetune_fraction_high=0.05,   # 5% of base
        notes="SFT/post-training phase of CogVideoX. FLOP drawn as "
              "base_flop * LogNormal(p10=0.1%,p90=5%) in MC. Global MFU (standard finetuning).",
    ),
    TrainingRunEstimate(
        model_name="AutoGLM-Rumination",
        org="zhipu",
        release_date="2025-04",
        training_flop_low=2.88e22 * 0.01,    # display: ≈ 2.88e20
        training_flop_high=2.88e23 * 0.10,   # display: ≈ 2.88e22
        epoch_estimate=False,
        mfu_override_low=0.01,  # RL-based agent -> very low MFU
        mfu_override_high=0.10,
        base_model_ref="GLM-Z1-Rumination-32B-0414",  # correlated with Z1 draw
        finetune_fraction_low=0.01,   # 1% of base
        finetune_fraction_high=0.10,  # 10% of base
        notes="Agent model. RL finetune on GLM-Z1. FLOP drawn as "
              "base_flop * LogNormal(p10=1%,p90=10%) in MC. "
              "LOW MFU: RL workload is inference-heavy.",
    ),
    # ---- Manually added models (not in CSV at all) ----
    TrainingRunEstimate(
        model_name="GLM-4-Voice (pretraining)",
        org="zhipu",
        release_date="2024-12",
        training_flop_low=4e22,
        training_flop_high=8e22,
        epoch_estimate=False,
        flop_distribution="uniform",
        notes="Voice model pretraining. Completely absent from Epoch CSV. "
              "Hardcoded p10=4e22, p90=8e22 based on manual assessment.",
    ),
    TrainingRunEstimate(
        model_name="GLM-4-Voice (SFT phase)",
        org="zhipu",
        release_date="2024-12",
        epoch_estimate=False,
        base_model_ref="GLM-4-Voice (pretraining)",
        finetune_fraction_low=0.001,   # 0.1% of base
        finetune_fraction_high=0.05,   # 5% of base
        notes="SFT/post-training phase of GLM-4-Voice. FLOP drawn as "
              "base_flop * LogNormal(p10=0.1%,p90=5%) in MC. Global MFU (standard finetuning).",
    ),
    TrainingRunEstimate(
        model_name="GLM-Realtime (pretraining)",
        org="zhipu",
        release_date="2025-01",
        training_flop_low=1.3e23,
        training_flop_high=3.2e23,
        epoch_estimate=False,
        flop_distribution="uniform",
        notes="Realtime multimodal model pretraining. Completely absent from Epoch CSV. "
              "Hardcoded p10=1.3e23, p90=3.2e23 based on manual assessment.",
    ),
    TrainingRunEstimate(
        model_name="GLM-Realtime (SFT phase)",
        org="zhipu",
        release_date="2025-01",
        epoch_estimate=False,
        base_model_ref="GLM-Realtime (pretraining)",
        finetune_fraction_low=0.001,   # 0.1% of base
        finetune_fraction_high=0.05,   # 5% of base
        notes="SFT/post-training phase of GLM-Realtime. FLOP drawn as "
              "base_flop * LogNormal(p10=0.1%,p90=5%) in MC. Global MFU (standard finetuning).",
    ),
]


# =============================================================================
# ADJUST BOUNDS: add estimated post-training for base pretraining models
# =============================================================================
# Base pretraining models (e.g. GLM-4.5, GLM-4.5-Air) have training_flop_low ==
# training_flop_high because Epoch only reports pretraining FLOP. But these models
# also underwent post-training (SFT, RLHF, etc.) which isn't captured.
#
# We estimate post-training (SFT + RLHF) as 1%-10% of base pretraining FLOP:
#   low  = base_flop * 1.01   (post-training adds at least 1%)
#   high = base_flop * 1.10   (post-training adds at most 10%)

_POST_TRAINING_LO = 0.01   # 1% of base
_POST_TRAINING_HI = 0.10   # 10% of base

_BASE_PRETRAINING_MODELS = {
    "MiniMax-Text-01", "GLM-4.5", "GLM-4.5-Air",
    "GLM-4-32B-0414", "GLM-4-9B-0414",
}

for _models in [MINIMAX_MODELS, ZHIPU_MODELS]:
    for _m in _models:
        if _m.model_name in _BASE_PRETRAINING_MODELS and _m.training_flop_low == _m.training_flop_high:
            base_flop = _m.training_flop_low
            _m.training_flop_low = base_flop * (1 + _POST_TRAINING_LO)
            _m.training_flop_high = base_flop * (1 + _POST_TRAINING_HI)

print(f"Post-training overhead applied to base model bounds: "
      f"{_POST_TRAINING_LO:.0%} (low) to {_POST_TRAINING_HI:.0%} (high)")



# =============================================================================
# AGGREGATION (Monte Carlo)
# =============================================================================
# All low/high values are treated as 10th and 90th percentiles of log-normal
# distributions. We sample:
#   - MFU:   once per draw (shared across all models — systematic uncertainty)
#   - Price: once per draw (shared — it's a market-level parameter)
#   - FLOP:  independently per model
#
# This means MFU and price are perfectly correlated across models within a
# draw, while FLOP uncertainty is independent per model.

N_SAMPLES = 50_000
RNG = np.random.default_rng(42)


def lognormal_from_percentiles(
    p10: float, p90: float, n: int, rng: np.random.Generator = RNG,
) -> np.ndarray:
    """Draw n samples from a log-normal where p10 and p90 are the 10th/90th percentiles.

    If p10 == p90, returns a constant array (point estimate, no uncertainty).
    """
    if p10 == p90:
        return np.full(n, p10)
    # z-score for the 90th percentile of a standard normal
    z90 = 1.2816  # scipy.stats.norm.ppf(0.9)
    log_p10 = np.log(p10)
    log_p90 = np.log(p90)
    mu = (log_p10 + log_p90) / 2
    sigma = (log_p90 - log_p10) / (2 * z90)
    return rng.lognormal(mu, sigma, n)


def compute_aggregate_training_cost_mc(
    models: list[TrainingRunEstimate],
    gpu_peak_flops: float = GPU_PEAK_FLOPS,
    mfu_p10: float = MFU_LOW,
    mfu_p90: float = MFU_HIGH,
    price_p10: float = COST_PER_GPU_HOUR_LOW,
    price_p90: float = COST_PER_GPU_HOUR_HIGH,
    n_samples: int = N_SAMPLES,
    rng: np.random.Generator = RNG,
) -> tuple[tuple[float, float, float], list[dict]]:
    """
    Monte Carlo estimate of aggregate final training run cost.

    Returns:
        (p5, p50, p95) aggregate totals in USD millions,
        list of per-model dicts with keys: model_name, org, release_date,
            flop_low, flop_high, epoch_estimate, cost_p5, cost_p50, cost_p95.
    """
    # Draw shared parameters (same across all models in each sample)
    mfu_samples_global = lognormal_from_percentiles(mfu_p10, mfu_p90, n_samples, rng)
    price_samples = lognormal_from_percentiles(price_p10, price_p90, n_samples, rng)

    # Precompute shared cost multiplier (for models using global MFU)
    # cost = flop / (peak * mfu) / 3600 * price
    cost_multiplier_global = price_samples / (gpu_peak_flops * mfu_samples_global * 3600)

    total_cost_samples = np.zeros(n_samples)
    per_model_results = []
    # Store drawn FLOP samples per model for correlated finetune references.
    # For in-list base models, keyed by model_name (stores the model's own FLOP).
    # For phantom bases (not in list), keyed by base_model_ref.
    model_flop_draws: dict[str, np.ndarray] = {}

    for m in models:
        # --- Determine FLOP samples ---
        if m.finetune_fraction_low is not None and m.finetune_fraction_high is not None:
            # Derived finetune model: FLOP = base_flop * fraction
            # This properly propagates uncertainty from both sources.
            if m.base_model_ref and m.base_model_ref in model_flop_draws:
                # Correlated draw: reuse base model's (or phantom base's) FLOP samples
                base_flop_samples = model_flop_draws[m.base_model_ref]
            elif m.base_flop_low is not None and m.base_flop_high is not None:
                # Phantom base: draw fresh and cache under base_model_ref so
                # sibling finetunes (e.g. I2V + T2V Directors) share the draw.
                if m.base_flop_distribution == "uniform":
                    base_flop_samples = rng.uniform(
                        m.base_flop_low, m.base_flop_high, n_samples)
                else:
                    base_flop_samples = lognormal_from_percentiles(
                        m.base_flop_low, m.base_flop_high, n_samples, rng)
                if m.base_model_ref:
                    model_flop_draws[m.base_model_ref] = base_flop_samples
            else:
                print(f"  {m.model_name:40s}  *** NO BASE FLOP SOURCE — skipped ***")
                continue
            frac_samples = lognormal_from_percentiles(
                m.finetune_fraction_low, m.finetune_fraction_high, n_samples, rng)
            flop_samples = base_flop_samples * frac_samples
        elif m.training_flop_low is not None and m.training_flop_high is not None:
            if m.flop_distribution == "uniform":
                flop_samples = rng.uniform(
                    m.training_flop_low, m.training_flop_high, n_samples,
                )
            else:
                flop_samples = lognormal_from_percentiles(
                    m.training_flop_low, m.training_flop_high, n_samples, rng,
                )
        else:
            print(f"  {m.model_name:40s}  *** MISSING — skipped ***")
            continue

        model_flop_draws[m.model_name] = flop_samples

        # Use per-model MFU if overridden (e.g. for RL-heavy models)
        model_gpu_peak = m.gpu_peak_flops_override if m.gpu_peak_flops_override is not None else gpu_peak_flops
        if m.mfu_override_low is not None and m.mfu_override_high is not None:
            mfu_samples_model = lognormal_from_percentiles(
                m.mfu_override_low, m.mfu_override_high, n_samples, rng,
            )
            cost_multiplier = price_samples / (model_gpu_peak * mfu_samples_model * 3600)
        elif m.gpu_peak_flops_override is not None:
            cost_multiplier = price_samples / (model_gpu_peak * mfu_samples_global * 3600)
        else:
            cost_multiplier = cost_multiplier_global

        model_cost = flop_samples * cost_multiplier  # USD
        total_cost_samples += model_cost

        # Report per-model percentiles (FLOP bounds from actual MC draws)
        flop_p5, flop_p95 = np.percentile(flop_samples, [5, 95])
        mp5, mp50, mp95 = np.percentile(model_cost / 1e6, [5, 50, 95])
        src = "Epoch" if m.epoch_estimate else "ROUGH GUESS"
        print(f"  {m.model_name:40s}  "
              f"p5=${mp5:7.2f}M  p50=${mp50:7.2f}M  p95=${mp95:7.2f}M  [{src}]")
        per_model_results.append({
            "model_name": m.model_name,
            "org": m.org,
            "release_date": m.release_date,
            "flop_low": flop_p5,
            "flop_high": flop_p95,
            "epoch_estimate": m.epoch_estimate,
            "cost_p5_usd_mn": mp5,
            "cost_p50_usd_mn": mp50,
            "cost_p95_usd_mn": mp95,
        })

    total_millions = total_cost_samples / 1e6
    p5, p50, p95 = np.percentile(total_millions, [5, 50, 95])
    return (p5, p50, p95), per_model_results



# =============================================================================
# OPENAI DATA (2024, from published analysis)
# =============================================================================
# Source: https://epoch.ai/data-insights/openai-compute-spend
OPENAI_FINAL_TRAINING_2024 = 480    # $480M — final training runs
OPENAI_OTHER_RD_2024 = 4520         # $4.52B — other R&D compute
OPENAI_RD_COMPUTE_2024 = 5000       # $5B — total R&D compute
OPENAI_INFERENCE_2024 = 2000        # $2B — inference compute
OPENAI_TOTAL_COMPUTE_2024 = 7000    # $7B — total compute (R&D + inference)

# OpenAI 90% CI breakdown (from Epoch's published analysis):
#   GPT-4.5 final training run: 90% CI = [$170M, $890M]
#   Other models (GPT-4o, GPT-4o mini, Sora Turbo, o-series post-training,
#     Q2 2024–Q1 2025): 90% CI = [$24M, $435M]
OPENAI_GPT45_TRAINING_LO = 170     # 5th percentile
OPENAI_GPT45_TRAINING_HI = 890     # 95th percentile
OPENAI_OTHER_TRAINING_LO = 24      # 5th percentile
OPENAI_OTHER_TRAINING_HI = 435     # 95th percentile
OPENAI_TRAINING_LO = OPENAI_GPT45_TRAINING_LO + OPENAI_OTHER_TRAINING_LO   # $194M
OPENAI_TRAINING_HI = OPENAI_GPT45_TRAINING_HI + OPENAI_OTHER_TRAINING_HI   # $1,325M


# =============================================================================
# Training Compute vs Total Compute (12-month windows from Excel)
# =============================================================================

def build_summary_and_plot(save_path: str | None = os.path.join(_OUTPUT_DIR, "training_vs_total_compute.png")):
    """
    Build a summary dataframe and plot comparing estimated training compute
    cost against total compute spending for MiniMax and Zhipu, using 12-month
    spending windows derived from the IPO Excel files.

    MiniMax window: Q4 2024 + Q1-Q3 2025
    Zhipu window:   H2 2024 + H1 2025
    """
    # --- Load financial data from Excel ---
    mm_fin = load_minimax_financials()
    zp_fin = load_zhipu_financials()

    print(f"\n{'='*70}")
    print("Financial data from Excel (12-month spending windows)")
    print(f"{'='*70}")
    print(f"MiniMax ({mm_fin['window']}):")
    print(f"  R&D compute:       ${mm_fin['rd_compute']:.1f}M")
    print(f"  Inference compute:  ${mm_fin['inference_compute']:.1f}M")
    print(f"  Total compute:      ${mm_fin['total_compute']:.1f}M")
    print(f"Zhipu ({zp_fin['window']}):")
    print(f"  R&D compute:       ${zp_fin['rd_compute']:.1f}M")
    print(f"  Inference compute:  ${zp_fin['inference_compute']:.1f}M")
    print(f"  Total compute:      ${zp_fin['total_compute']:.1f}M")
    if "rmb_per_usd" in zp_fin:
        print(f"  (RMB/USD rate: {zp_fin['rmb_per_usd']:.3f})")

    # --- Monte Carlo training cost estimates ---
    # Use a fresh RNG so results are reproducible regardless of what ran before
    rng = np.random.default_rng(42)

    print(f"\n{'='*70}")
    print("Monte Carlo training cost estimates")
    print(f"{'='*70}")

    print("\n--- MiniMax ---")
    (mm_p5, mm_p50, mm_p95), mm_per_model = compute_aggregate_training_cost_mc(
        MINIMAX_MODELS, rng=rng,
    )
    print(f"  {'TOTAL':40s}  "
          f"p5=${mm_p5:7.2f}M  p50=${mm_p50:7.2f}M  p95=${mm_p95:7.2f}M")

    print("\n--- Zhipu ---")
    (zp_p5, zp_p50, zp_p95), zp_per_model = compute_aggregate_training_cost_mc(
        ZHIPU_MODELS, rng=rng,
    )
    print(f"  {'TOTAL':40s}  "
          f"p5=${zp_p5:7.2f}M  p50=${zp_p50:7.2f}M  p95=${zp_p95:7.2f}M")

    # --- Build summary dataframe ---
    # Compare training compute cost against R&D compute from IPO filings
    # (R&D compute is the most relevant denominator since it excludes inference)
    summary = pd.DataFrame([
        {
            "company": "MiniMax",
            "spending_window": mm_fin["window"],
            "training_compute_p5": mm_p5,
            "training_compute_p50": mm_p50,
            "training_compute_p95": mm_p95,
            "rd_compute": mm_fin["rd_compute"],
            "total_compute": mm_fin["total_compute"],
            "ratio_p5_rd": mm_p5 / mm_fin["rd_compute"],
            "ratio_p50_rd": mm_p50 / mm_fin["rd_compute"],
            "ratio_p95_rd": mm_p95 / mm_fin["rd_compute"],
            "ratio_p5_total": mm_p5 / mm_fin["total_compute"],
            "ratio_p50_total": mm_p50 / mm_fin["total_compute"],
            "ratio_p95_total": mm_p95 / mm_fin["total_compute"],
        },
        {
            "company": "Z.ai / Zhipu",
            "spending_window": zp_fin["window"],
            "training_compute_p5": zp_p5,
            "training_compute_p50": zp_p50,
            "training_compute_p95": zp_p95,
            "rd_compute": zp_fin["rd_compute"],
            "total_compute": zp_fin["total_compute"],
            "ratio_p5_rd": zp_p5 / zp_fin["rd_compute"],
            "ratio_p50_rd": zp_p50 / zp_fin["rd_compute"],
            "ratio_p95_rd": zp_p95 / zp_fin["rd_compute"],
            "ratio_p5_total": zp_p5 / zp_fin["total_compute"],
            "ratio_p50_total": zp_p50 / zp_fin["total_compute"],
            "ratio_p95_total": zp_p95 / zp_fin["total_compute"],
        },
    ])

    print(f"\n{'='*70}")
    print("Summary Table")
    print(f"{'='*70}")
    with pd.option_context("display.float_format", "{:.2f}".format, "display.width", 120):
        print(summary.to_string(index=False))

    # --- Per-model training cost table (Excel) ---
    per_model_df = pd.DataFrame(mm_per_model + zp_per_model)
    per_model_df = per_model_df.rename(columns={
        "model_name": "Model",
        "org": "Company",
        "release_date": "Release date",
        "flop_low": "Training FLOP (p5)",
        "flop_high": "Training FLOP (p95)",
        "epoch_estimate": "Epoch estimate?",
        "cost_p5_usd_mn": "Training cost p5 ($M)",
        "cost_p50_usd_mn": "Training cost p50 ($M)",
        "cost_p95_usd_mn": "Training cost p95 ($M)",
    })
    per_model_path = os.path.join(_OUTPUT_DIR, "per_model_training_cost.xlsx")
    per_model_df.to_excel(per_model_path, index=False, float_format="%.2f")
    print(f"\nSaved per-model training cost table to {per_model_path}")

    # --- Plot: grouped bar chart ---
    fig, ax = plt.subplots(figsize=(11, 7))

    firms = ["MiniMax", "Z.ai / Zhipu"]
    x = np.arange(len(firms))
    bar_width = 0.22

    training_p50 = [mm_p50, zp_p50]
    training_err_lo = [mm_p50 - mm_p5, zp_p50 - zp_p5]
    training_err_hi = [mm_p95 - mm_p50, zp_p95 - zp_p50]
    rd_vals = [mm_fin["rd_compute"], zp_fin["rd_compute"]]

    # Training cost bars with error bars
    bars_train = ax.bar(
        x - bar_width / 2,
        training_p50,
        bar_width,
        yerr=[training_err_lo, training_err_hi],
        capsize=5,
        label="Est. final training cost\n(median, 90% CI)",
        color="#2ca02c",
        edgecolor="white",
        zorder=3,
    )

    # R&D compute bars
    bars_rd = ax.bar(
        x + bar_width / 2,
        rd_vals,
        bar_width,
        label="R&D compute (IPO filings)",
        color="#5b7fbf",
        edgecolor="white",
        zorder=3,
    )

    # Label training bars with median and CI
    for i, bar in enumerate(bars_train):
        p5_v = mm_p5 if i == 0 else zp_p5
        p50_v = mm_p50 if i == 0 else zp_p50
        p95_v = mm_p95 if i == 0 else zp_p95
        top = p95_v + max(5, p95_v * 0.06)
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            top,
            f"${p50_v:.0f}M\n[${p5_v:.0f}M-${p95_v:.0f}M]",
            ha="center", va="bottom",
            fontsize=8, fontweight="bold",
        )

    # Label R&D bars
    for bar in bars_rd:
        h = bar.get_height()
        if h > 0:
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                h + max(3, h * 0.03),
                f"${h:.0f}M",
                ha="center", va="bottom",
                fontsize=9, fontweight="bold",
            )

    # Add ratio annotations below x-axis labels
    for i, firm in enumerate(firms):
        rd_ratio = training_p50[i] / rd_vals[i]
        ax.text(
            x[i], -0.12, f"Training/R&D: {rd_ratio:.0%}",
            ha="center", va="top", fontsize=8, color="gray",
            transform=ax.get_xaxis_transform(),
        )

    ax.set_ylabel("USD millions", fontsize=12)
    ax.set_title(
        "Estimated Final Training Run Cost vs. Compute Spending\n"
        "(MiniMax: Q4 2024-Q3 2025; Zhipu: H2 2024-H1 2025)",
        fontsize=12,
        fontweight="bold",
    )
    ax.set_xticks(x)
    ax.set_xticklabels(firms, fontsize=12)
    ax.legend(loc="upper left", fontsize=9)
    ax.grid(axis="y", alpha=0.3, zorder=0)
    ax.set_axisbelow(True)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"${v:,.0f}M"))

    # Leave room at bottom for ratio text
    ax.set_xlim(-0.6, len(firms) - 0.4)
    plt.subplots_adjust(bottom=0.18)

    fig.text(0.98, 0.02, "Epoch AI", ha="right", fontsize=9, color="gray", style="italic")

    plt.tight_layout()

    if save_path:
        fig.savefig(save_path, dpi=200, bbox_inches="tight")
        print(f"\nSaved plot to {save_path}")

    plt.show()

    # --- Marimekko charts ---
    marimekko_args = dict(
        mm_p5=mm_p5, mm_p50=mm_p50, mm_p95=mm_p95,
        zp_p5=zp_p5, zp_p50=zp_p50, zp_p95=zp_p95,
        mm_rd=mm_fin["rd_compute"], zp_rd=zp_fin["rd_compute"],
    )
    plot_marimekko(**marimekko_args)
    plot_marimekko_with_openai(**marimekko_args)

    # --- No-Hailuo variant (exclude Hailuo models from MiniMax) ---
    mm_no_hailuo = [m for m in MINIMAX_MODELS if "Hailuo" not in m.model_name]
    rng_nh = np.random.default_rng(42)
    (mm_nh_p5, mm_nh_p50, mm_nh_p95), _ = compute_aggregate_training_cost_mc(
        mm_no_hailuo, rng=rng_nh,
    )
    print(f"\n--- MiniMax (no Hailuo) ---")
    print(f"  {'TOTAL':40s}  "
          f"p5=${mm_nh_p5:7.2f}M  p50=${mm_nh_p50:7.2f}M  p95=${mm_nh_p95:7.2f}M")
    marimekko_no_hailuo_args = dict(
        mm_p5=mm_nh_p5, mm_p50=mm_nh_p50, mm_p95=mm_nh_p95,
        zp_p5=zp_p5, zp_p50=zp_p50, zp_p95=zp_p95,
        mm_rd=mm_fin["rd_compute"], zp_rd=zp_fin["rd_compute"],
    )
    plot_marimekko(**marimekko_no_hailuo_args,
                   save_path=os.path.join(_OUTPUT_DIR, "marimekko_training_compute_no_hailuo.html"))
    plot_marimekko_with_openai(**marimekko_no_hailuo_args,
                               save_path=os.path.join(_OUTPUT_DIR, "marimekko_compute_openai_no_hailuo.html"))

    return summary


# =============================================================================
# PSEUDO-MARIMEKKO: Training compute breakdown with uncertainty (interactive)
# =============================================================================

def plot_marimekko(
    mm_p5: float, mm_p50: float, mm_p95: float,
    zp_p5: float, zp_p50: float, zp_p95: float,
    mm_rd: float, zp_rd: float,
    save_path: str | None = os.path.join(_OUTPUT_DIR, "marimekko_training_compute.html"),
):
    """
    Interactive pseudo-Marimekko chart (plotly).

    Each company gets one horizontal bar (0–100% of R&D compute).
    Segments:
      [0, p50]   — pink       (median training compute)
      [p50, rd]  — muted teal (other R&D compute)

    A horizontal error bar (p5–p95, 90% CI) overlays the training segment
    to show uncertainty. Hover on any segment shows both % and USD.
    """
    import plotly.graph_objects as go

    PINK = "#d6627e"
    TEAL = "#76b7b2"
    ERR_COLOR = "#8b0000"
    BG_COLOR = "#f7f5f0"
    GRID_COLOR = "#d4d0c8"

    companies = ["MiniMax", "Z.ai / Zhipu"]
    p5s = [mm_p5, zp_p5]
    p50s = [mm_p50, zp_p50]
    p95s = [mm_p95, zp_p95]
    rds = [mm_rd, zp_rd]

    fig = go.Figure()

    # --- Build 2 segments per company (training + other R&D) ---
    seg_defs = []
    for i in range(len(companies)):
        p50, rd = p50s[i], rds[i]
        other_rd = max(0, rd - p50)
        seg_defs.append([
            (p50 / rd * 100, p50,
             PINK, "Training compute (median)"),
            (other_rd / rd * 100, other_rd,
             TEAL, "Other R&D compute"),
        ])

    segment_labels = [s[3] for s in seg_defs[0]]
    segment_colors = [s[2] for s in seg_defs[0]]

    for seg_idx in range(2):
        widths = [seg_defs[i][seg_idx][0] for i in range(len(companies))]
        usds = [seg_defs[i][seg_idx][1] for i in range(len(companies))]
        bases = [sum(seg_defs[i][j][0] for j in range(seg_idx))
                 for i in range(len(companies))]

        hover_texts = [
            f"<b>{companies[i]}</b><br>"
            f"{segment_labels[seg_idx]}<br>"
            f"${usds[i]:.1f}M  ({widths[i]:.1f}%)"
            for i in range(len(companies))
        ]

        fig.add_trace(go.Bar(
            y=companies,
            x=widths,
            base=bases,
            orientation="h",
            name=segment_labels[seg_idx],
            marker_color=segment_colors[seg_idx],
            marker_line=dict(color="white", width=1),
            hovertext=hover_texts,
            hoverinfo="text",
        ))

    # --- Error bar trace (p5–p95, 90% CI) ---
    p50_pcts = [p50s[i] / rds[i] * 100 for i in range(len(companies))]
    err_lo = [(p50s[i] - p5s[i]) / rds[i] * 100 for i in range(len(companies))]
    err_hi = [(p95s[i] - p50s[i]) / rds[i] * 100 for i in range(len(companies))]

    hover_err = [
        f"<b>{companies[i]}</b><br>"
        f"Training compute 90% CI<br>"
        f"p5: ${p5s[i]:.1f}M  ({p5s[i]/rds[i]*100:.1f}%)<br>"
        f"p50: ${p50s[i]:.1f}M  ({p50_pcts[i]:.1f}%)<br>"
        f"p95: ${p95s[i]:.1f}M  ({p95s[i]/rds[i]*100:.1f}%)"
        for i in range(len(companies))
    ]

    fig.add_trace(go.Scatter(
        y=companies,
        x=p50_pcts,
        mode="markers",
        marker=dict(symbol="line-ns", size=12, color=ERR_COLOR, line_width=2),
        error_x=dict(
            type="data",
            symmetric=False,
            array=err_hi,
            arrayminus=err_lo,
            color=ERR_COLOR,
            thickness=2,
            width=8,
        ),
        name="Training compute (90% CI)",
        hovertext=hover_err,
        hoverinfo="text",
    ))

    fig.update_layout(
        barmode="overlay",
        plot_bgcolor=BG_COLOR,
        paper_bgcolor=BG_COLOR,
        font=dict(family="Arial, sans-serif", color="#333333"),
        title=dict(
            text="Estimated Training Compute as Share of R&D Compute Spending",
            font=dict(size=15),
        ),
        xaxis=dict(
            title="% of R&D compute spending",
            ticksuffix="%",
            range=[0, 105],
            gridcolor=GRID_COLOR,
            zeroline=False,
        ),
        yaxis=dict(
            autorange="reversed",
            tickfont=dict(size=13),
        ),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=-0.35,
            xanchor="center",
            x=0.5,
            font=dict(size=11),
        ),
        height=320,
        width=900,
        margin=dict(l=120, r=40, t=60, b=90),
        annotations=[
            dict(
                text="Epoch AI",
                xref="paper", yref="paper",
                x=1.0, y=-0.45,
                showarrow=False,
                font=dict(size=10, color="gray"),
            ),
        ],
    )

    if save_path:
        fig.write_html(save_path)
        print(f"\nSaved Marimekko chart to {save_path}")

    fig.show()


def plot_marimekko_with_openai(
    mm_p5: float, mm_p50: float, mm_p95: float,
    zp_p5: float, zp_p50: float, zp_p95: float,
    mm_rd: float, zp_rd: float,
    save_path: str | None = os.path.join(_OUTPUT_DIR, "marimekko_compute_openai.html"),
):
    """
    Same Marimekko chart as plot_marimekko but with OpenAI added as a third
    bar. All three companies have 90% CI error bars.
    R&D compute is the denominator (inference excluded).
    """
    import plotly.graph_objects as go

    PINK = "#d6627e"
    TEAL = "#76b7b2"
    ERR_COLOR = "#8b0000"
    BG_COLOR = "#f7f5f0"
    GRID_COLOR = "#d4d0c8"

    companies = ["OpenAI", "MiniMax", "Z.ai / Zhipu"]
    training_p50 = [OPENAI_FINAL_TRAINING_2024, mm_p50, zp_p50]
    rds = [OPENAI_RD_COMPUTE_2024, mm_rd, zp_rd]

    fig = go.Figure()

    # --- Build 2 segments per company (training + other R&D) ---
    seg_defs = []
    for i in range(len(companies)):
        p50, rd = training_p50[i], rds[i]
        other_rd = max(0, rd - p50)
        seg_defs.append([
            (p50 / rd * 100, p50, PINK, "Training compute (median)"),
            (other_rd / rd * 100, other_rd, TEAL, "Other R&D compute"),
        ])

    segment_labels = [s[3] for s in seg_defs[0]]
    segment_colors = [s[2] for s in seg_defs[0]]

    for seg_idx in range(2):
        widths = [seg_defs[i][seg_idx][0] for i in range(len(companies))]
        usds = [seg_defs[i][seg_idx][1] for i in range(len(companies))]
        bases = [sum(seg_defs[i][j][0] for j in range(seg_idx))
                 for i in range(len(companies))]

        hover_texts = [
            f"<b>{companies[i]}</b><br>"
            f"{segment_labels[seg_idx]}<br>"
            f"${usds[i]:,.1f}M  ({widths[i]:.1f}%)"
            for i in range(len(companies))
        ]

        fig.add_trace(go.Bar(
            y=companies,
            x=widths,
            base=bases,
            orientation="h",
            name=segment_labels[seg_idx],
            marker_color=segment_colors[seg_idx],
            marker_line=dict(color="white", width=1),
            hovertext=hover_texts,
            hoverinfo="text",
        ))

    # --- Error bars (90% CI) for all three companies ---
    all_p5s = [OPENAI_TRAINING_LO, mm_p5, zp_p5]
    all_p50s = [OPENAI_FINAL_TRAINING_2024, mm_p50, zp_p50]
    all_p95s = [OPENAI_TRAINING_HI, mm_p95, zp_p95]

    p50_pcts = [all_p50s[i] / rds[i] * 100 for i in range(3)]
    err_lo = [(all_p50s[i] - all_p5s[i]) / rds[i] * 100 for i in range(3)]
    err_hi = [(all_p95s[i] - all_p50s[i]) / rds[i] * 100 for i in range(3)]

    hover_err = [
        f"<b>{companies[i]}</b><br>"
        f"Training compute 90% CI<br>"
        f"p5: ${all_p5s[i]:,.1f}M  ({all_p5s[i]/rds[i]*100:.1f}%)<br>"
        f"p50: ${all_p50s[i]:,.1f}M  ({p50_pcts[i]:.1f}%)<br>"
        f"p95: ${all_p95s[i]:,.1f}M  ({all_p95s[i]/rds[i]*100:.1f}%)"
        for i in range(3)
    ]

    fig.add_trace(go.Scatter(
        y=companies,
        x=p50_pcts,
        mode="markers",
        marker=dict(symbol="line-ns", size=12, color=ERR_COLOR, line_width=2),
        error_x=dict(
            type="data",
            symmetric=False,
            array=err_hi,
            arrayminus=err_lo,
            color=ERR_COLOR,
            thickness=2,
            width=8,
        ),
        name="Training compute (90% CI)",
        hovertext=hover_err,
        hoverinfo="text",
    ))

    fig.update_layout(
        barmode="overlay",
        plot_bgcolor=BG_COLOR,
        paper_bgcolor=BG_COLOR,
        font=dict(family="Arial, sans-serif", color="#333333"),
        title=dict(
            text="Estimated Training Compute as Share of R&D Compute Spending",
            font=dict(size=15),
        ),
        xaxis=dict(
            title="% of R&D compute spending",
            ticksuffix="%",
            range=[0, 105],
            gridcolor=GRID_COLOR,
            zeroline=False,
        ),
        yaxis=dict(
            autorange="reversed",
            tickfont=dict(size=13),
        ),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=-0.3,
            xanchor="center",
            x=0.5,
            font=dict(size=11),
        ),
        height=380,
        width=900,
        margin=dict(l=120, r=40, t=60, b=90),
        annotations=[
            dict(
                text="Epoch AI",
                xref="paper", yref="paper",
                x=1.0, y=-0.38,
                showarrow=False,
                font=dict(size=10, color="gray"),
            ),
        ],
    )

    if save_path:
        fig.write_html(save_path)
        print(f"\nSaved Marimekko chart (with OpenAI) to {save_path}")

    fig.show()


# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":
    print("=" * 70)
    print("R&D Compute vs. Final Training Run Compute")
    print("=" * 70)

    print(f"\nConversion parameters:")
    print(f"  GPU: H800 dense BF16 peak = {GPU_PEAK_FLOPS:.3e} FLOP/s")
    print(f"  MFU range: {MFU_LOW} (pessimistic) - {MFU_HIGH} (optimistic)")
    print(f"  GPU-hour price range: ${COST_PER_GPU_HOUR_LOW} - ${COST_PER_GPU_HOUR_HIGH}")

    # Count models with/without Epoch estimates
    for label, models in [("MiniMax", MINIMAX_MODELS), ("Zhipu", ZHIPU_MODELS)]:
        n_epoch = sum(1 for m in models if m.epoch_estimate)
        n_rough = sum(1 for m in models if not m.epoch_estimate)
        n_total = len(models)
        print(f"\n{label}: {n_total} models total — "
              f"{n_epoch} with Epoch FLOP, {n_rough} with regression/manual bounds")
        for m in models:
            src = "Epoch" if m.epoch_estimate else "est."
            flop_lo = f"{m.training_flop_low:.2e}" if m.training_flop_low is not None else "derived"
            flop_hi = f"{m.training_flop_high:.2e}" if m.training_flop_high is not None else "derived"
            print(f"  {m.model_name:40s}  {m.release_date}  "
                  f"FLOP=[{flop_lo}, {flop_hi}]  [{src}]")

    # Build summary table and new plot (12-month windows from Excel)
    summary = build_summary_and_plot(save_path=os.path.join(_OUTPUT_DIR, "training_vs_total_compute.png"))
